import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures

file = load_workbook("./resource/base_excel.xlsx")  # 엑셀 파일 위치
worksheets = ['2020년 시총500억미만', '2020년 시총1000억미만', '2020년 시총2000억미만']  # 워크시트

def get_stock_id(sheet):
    '''
    엑셀 파일을 읽는다.
    :param sheet: worksheet 이름
    :return: [주식번호, 상장주식수, 현재가, 52주, 매출액2018, 매출액2019, 영업이익2018, 영업이익2019]
    '''

    stock_list = []
    start = 'D3'
    end = 'L' + str(sheet.max_row)
    for row in sheet[start: end]:
        v = row[0].value
        if v is not None:
            col = [v]
            for i in range(2, 9):
                col.append(row[i].value)
            stock_list.append(col)
    return stock_list


def string_to_int(s):
    try:
        i = int(s.replace(",", ""))
    except ValueError:
        i = 0
    return i


def get_company_information(stock_id, income=False):
    '''
    crawling
    :param stock_id: 주식번호
    :param income: True일 경우, 매출액2018, 매출액2019, 영업이익2018, 영업이익2019를 추가해서 반환
    :return: [상장주식수, 현재가, 52주 {, 매출액2018, 매출액2019, 영업이익2018, 영업이익2019}]
    '''

    str_url = "https://finance.naver.com/item/main.nhn?code=" + stock_id
    res = requests.get(str_url)
    soup = BeautifulSoup(res.text, 'html.parser')  # 파싱

    company_informations = []

    invest_information = soup.find(id="tab_con1").find_all("tr")  # 투자정보
    # 시가총액
    # amount = invest_information[0].find("td").get_text().split()
    # stock_informations.append(amount[0] + " " + amount[1])

    # 상장주식수
    company_informations.append(string_to_int(invest_information[2].find("td").get_text()))

    # 현재가
    company_informations.append(string_to_int(soup.find("p", "no_today").find("span").get_text()))

    # 52주 최저
    company_informations.append(string_to_int(invest_information[8].find_all("em")[1].get_text()))

    if income:
        # 매출액2018, 2019
        finacial_table = soup.find("table", "tb_type1 tb_num tb_type1_ifrs").find_all("tr")  # 기업실적분석
        revenue = finacial_table[3].find_all("td")
        for i in range(1, 3):
            _rev = string_to_int(revenue[i].get_text())
            company_informations.append(_rev)

        # 영업이익2018, 2019
        income = finacial_table[4].find_all("td")
        for i in range(1, 3):
            _inc = string_to_int(income[i].get_text())
            company_informations.append(_inc)

    return company_informations


def get_time():
    from time import localtime
    now = localtime()
    return "{}{}{}_{}{}{}".format(now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)

def multi_threading(worksheet):
    print(worksheet, "start")
    sheet = file[worksheet]
    stock_list = get_stock_id(sheet)
    stock_length = len(stock_list)
    for i in range(stock_length):
        company_infos = get_company_information(stock_list[i][0], income=True)
        for j in range(len(company_infos)):  # 갱신
            if stock_list[i][j + 1] != company_infos[j]:
                sheet.cell(row=i + 3, column=6 + j).value = company_infos[j]
    print(worksheet, "end")

if __name__ == '__main__':
    startTime = time.time()
    thread_list = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        for worksheet in worksheets:
            thread_list.append(executor.submit(multi_threading, worksheet))
        for executor in concurrent.futures.as_completed(thread_list):
            executor.result()
    file.save("./resource/result_" + get_time() + ".xlsx")  # 파일 저장
    print(time.time()-startTime)
